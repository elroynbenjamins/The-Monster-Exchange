import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path(r"C:\Users\elroy\Downloads\Monster Exchange Master Design Database v47 - 244 Species Release QA & Combat Integration.xlsx")
OUTPUT = Path("data/design-v47")
SHEETS = {
    "Species Catalog": "species-catalog.json",
    "Species Combat Stats": "species-combat-stats.json",
    "Species Passives": "species-passives.json",
    "Species Skill Learnsets": "species-skill-learnsets.json",
    "Species Trait Pools": "species-trait-pools.json",
    "Evolution Paths": "evolution-paths.json",
    "Monsterdex Catalog": "monsterdex-catalog.json",
    "Full Dex Order v2": "full-dex-order.json",
    "Species Ecology": "species-ecology.json",
    "Obtainability Matrix": "obtainability-matrix.json",
    "Species Market Baselines": "species-market-baselines.json",
    "Monster Visual Briefs": "monster-visual-briefs.json",
    "Monster Art Assets": "monster-art-assets.json",
    "Trait Catalog": "trait-catalog.json",
    "Skill Catalog": "skill-catalog.json",
    "Monsterdex Rules": "monsterdex-rules.json",
    "Starter Integration Matrix": "starter-integration-matrix.json",
}

book = load_workbook(SOURCE, read_only=True, data_only=True)
OUTPUT.mkdir(parents=True, exist_ok=True)
counts = {}
for sheet_name, filename in SHEETS.items():
    rows = list(book[sheet_name].iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else f"Column {index + 1}" for index, value in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if not any(value is not None and value != "" for value in row):
            continue
        records.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers)})
    (OUTPUT / filename).write_text(json.dumps(records, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    counts[sheet_name] = len(records)
print(json.dumps(counts, indent=2))
